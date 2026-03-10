from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app.routes.auth import login_required, role_required
from app.models.etudiant import (
    get_etudiants, get_etudiant_by_id, get_filieres,
    get_inscriptions_etudiant, get_notes_etudiant,
    create_etudiant, update_etudiant, delete_etudiant,
    get_all_etudiants_export,
)
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Response

etudiants_bp = Blueprint("etudiants", __name__, url_prefix="/etudiants")


@etudiants_bp.route("/")
@role_required("admin")
def index():
    page       = request.args.get("page", 1, type=int)
    search     = request.args.get("search", "")
    filiere_id = request.args.get("filiere_id", "")
    filieres   = get_filieres()
    etudiants, total, total_pages = get_etudiants(page, search, filiere_id)
    return render_template("etudiants/index.html",
        etudiants=etudiants, total=total, total_pages=total_pages,
        page=page, search=search, filiere_id=filiere_id, filieres=filieres)


@etudiants_bp.route("/<int:etudiant_id>")
@login_required
def detail(etudiant_id):
    # Étudiant ne peut voir que sa propre fiche
    if session.get("user_role") == "etudiant":
        if session.get("user_profile_id") != etudiant_id:
            flash("Accès refusé.", "danger")
            return redirect(url_for("dashboard.index"))
    etudiant     = get_etudiant_by_id(etudiant_id)
    if not etudiant:
        flash("Étudiant introuvable.", "danger")
        return redirect(url_for("dashboard.index"))
    inscriptions = get_inscriptions_etudiant(etudiant_id)
    notes        = get_notes_etudiant(etudiant_id)
    return render_template("etudiants/detail.html",
        etudiant=etudiant, inscriptions=inscriptions, notes=notes)


@etudiants_bp.route("/ajouter", methods=["GET", "POST"])
@role_required("admin")
def ajouter():
    if request.method == "POST":
        matricule       = request.form.get("matricule", "").strip()
        nom             = request.form.get("nom", "").strip()
        prenom          = request.form.get("prenom", "").strip()
        date_naissance  = request.form.get("date_naissance", "").strip()
        email           = request.form.get("email", "").strip()
        mot_de_passe    = request.form.get("mot_de_passe", "").strip()
        if not all([matricule, nom, prenom, date_naissance, email, mot_de_passe]):
            flash("Tous les champs sont obligatoires.", "danger")
            return render_template("etudiants/form.html", etudiant=None, action="Ajouter")
        try:
            create_etudiant(matricule, nom, prenom, date_naissance, email, mot_de_passe)
            flash(f"Étudiant {nom} {prenom} ajouté.", "success")
            return redirect(url_for("etudiants.index"))
        except Exception as e:
            flash(f"Erreur : {e}", "danger")
    return render_template("etudiants/form.html", etudiant=None, action="Ajouter")


@etudiants_bp.route("/modifier/<int:etudiant_id>", methods=["GET", "POST"])
@role_required("admin")
def modifier(etudiant_id):
    etudiant = get_etudiant_by_id(etudiant_id)
    if not etudiant:
        flash("Étudiant introuvable.", "danger")
        return redirect(url_for("etudiants.index"))
    if request.method == "POST":
        matricule      = request.form.get("matricule", "").strip()
        nom            = request.form.get("nom", "").strip()
        prenom         = request.form.get("prenom", "").strip()
        date_naissance = request.form.get("date_naissance", "").strip()
        email          = request.form.get("email", "").strip()
        try:
            update_etudiant(etudiant_id, matricule, nom, prenom, date_naissance, email, etudiant["id_utilisateur"])
            flash("Étudiant mis à jour.", "success")
            return redirect(url_for("etudiants.detail", etudiant_id=etudiant_id))
        except Exception as e:
            flash(f"Erreur : {e}", "danger")
    return render_template("etudiants/form.html", etudiant=etudiant, action="Modifier")


@etudiants_bp.route("/supprimer/<int:etudiant_id>", methods=["POST"])
@role_required("admin")
def supprimer(etudiant_id):
    try:
        delete_etudiant(etudiant_id)
        flash("Étudiant supprimé.", "success")
    except Exception as e:
        flash(f"Erreur : {e}", "danger")
    return redirect(url_for("etudiants.index"))


@etudiants_bp.route("/export/excel")
@role_required("admin")
def export_excel():
    rows = get_all_etudiants_export()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Étudiants"
    headers = ["Matricule","Nom","Prénom","Date naissance","Email","Classe"]
    hfill = PatternFill("solid", fgColor="0b0f1a"); hfont = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    rfill = PatternFill("solid", fgColor="f0f4ff")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            if ri % 2 == 0: c.fill = rfill
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=etudiants.xlsx"})
